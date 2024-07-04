import json

from django.core.files.storage import FileSystemStorage

import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook, load_workbook, worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date
from ..models import Products, Contact
from authentication.models import Employee


def main():
    print('Hewlo')






class PdfToExcel:

    def create_workbook(self, index: int, table: dict, path_excel: str, data):

        buyer_id = data['buyer']
        selling_date = data['sellingDate']
        selling_place = data['sellingPlace']
        car_plate = data['car_plate']
        incoterms = data['incoterms']
        incoterms_place = data['incoterms_place']
        payment_time = data['payment_time']
        notes = data['notes']
        username = data['username']

        excel_path = path_excel + 'Faktura ' + str(index) + '-' + date.today().strftime('%d-%m-%y') + '.xlsx'
        workbook = Workbook()
        workbook.create_sheet('Faktura')
        workbook.create_sheet('Specyfikacja')
        del workbook['Sheet']
        self.populate_seller(workbook['Faktura'],
                             car_plate=car_plate,
                             selling_date=selling_date,
                             selling_place=selling_place,
                             username=username
                             )
        self.populate_seller(workbook['Specyfikacja'],
                             car_plate=car_plate,
                             selling_date=selling_date,
                             selling_place=selling_place,
                             username=username
                             )
        self.populate_buyer(workbook['Faktura'],
                            buyer_id=buyer_id,
                            incoterms=incoterms,
                            incoterms_place=incoterms_place,
                            payment_time=payment_time,
                            notes=notes
                            )
        self.populate_buyer(workbook['Specyfikacja'],
                            buyer_id=buyer_id,
                            incoterms=incoterms,
                            incoterms_place=incoterms_place,
                            payment_time=payment_time,
                            notes=notes,
                            )
        sheet = workbook['Faktura']
        sheet['A1'] = "Faktura VAT"
        sheet['C1'] = '№ ' + str(index) + '-' + date.today().strftime('%d-%m/%y')
        sheet['I1'] = 'Poznan ' + date.today().strftime('%d.%m.%y')
        self.stylyse_workbook(workbook)
        workbook.save(excel_path)
        table = pd.DataFrame(table)
        (invoice_len, specification_len) = self.create_data_frames_and_save_to_excel_document(table, excel_path)
        self.add_footer(excel_path, invoice_len, specification_len)
        return excel_path

    def stylyse_workbook(self, workbook: Workbook):
        sheet = workbook['Faktura']
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 17
        sheet.column_dimensions['D'].width = 17
        sheet = workbook['Specyfikacja']
        sheet.column_dimensions['A'].width = 8
        sheet.column_dimensions['B'].width = 17
        sheet.column_dimensions['C'].width = 17
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 12
        sheet.column_dimensions['F'].width = 12
        sheet.column_dimensions['G'].width = 12

    def populate_seller(self, sheet: worksheet, car_plate: str, selling_date: str, selling_place: str, username: str):
        userdata = Employee.objects.get(user__username=username)
        sheet.merge_cells('A4:C4')
        sheet['A4'] = 'Sprzedający:'
        sheet.merge_cells('A5:C5')
        sheet['A5'] = userdata.company
        sheet.merge_cells('A6:C6')
        sheet['A6'] = userdata.name
        sheet.merge_cells('A7:C7')
        sheet['A7'] = userdata.street_address
        sheet.merge_cells('A8:C8')
        sheet['A8'] = userdata.postal_code + ' ' + userdata.city + ', ' + userdata.country
        sheet.merge_cells('A9:C9')
        sheet['A9'] = 'Nip:' + userdata.nip
        sheet.merge_cells('A10:C10')
        sheet['A10'] = userdata.bank_name
        sheet.merge_cells('A11:C11')
        sheet['A11'] = userdata.bank_address
        sheet.merge_cells('A12:C12')
        sheet['A12'] = userdata.iban
        sheet.merge_cells('A13:C13')
        sheet['A13'] = 'Swift: ' + userdata.swift
        sheet.merge_cells('A15:C15')
        sheet['A15'] = 'Data sprzedaży: ' + str(selling_date)
        sheet.merge_cells('A16:C16')
        sheet['A16'] = 'Miejsce załadunku :' + str(selling_place)
        sheet.merge_cells('A17:C17')
        sheet['A17'] = 'Samochód: ' + str(car_plate)

    def populate_buyer(self, sheet: worksheet, buyer_id, incoterms, incoterms_place, payment_time, notes):

        b_data = Contact.objects.get(id=buyer_id)

        sheet['I4'] = 'Kupujący: '
        sheet.merge_cells('G5:K5')
        sheet['G5'] = b_data.name
        sheet.merge_cells('G6:K6')
        sheet['G6'] = b_data.street
        sheet.merge_cells('G7:K7')
        sheet['G7'] = (b_data.postal_code or '' +
                       ' ' + b_data.city or '' +
                       ', ' + b_data.voivodeship or '' +
                       ', ' + b_data.country or '')
        sheet.merge_cells('G8:K8')
        sheet['G8'] = b_data.iban
        sheet.merge_cells('G9:K9')
        sheet['G9'] = 'SWIFT: '
        sheet['I10'] = 'Odbiorca:'
        sheet.merge_cells('G11:K11')
        sheet['G11'] = b_data.name
        sheet.merge_cells('G12:K12')
        sheet['G12'] = b_data.street
        sheet.merge_cells('G13:K13')
        sheet['G13'] = (b_data.postal_code or '' +
                        ' ' + b_data.city or '' +
                        ', ' + b_data.voivodeship or '' +
                        ', ' + b_data.country or '')
        sheet.merge_cells('G15:K15')
        sheet['G15'] = 'Warunki dostawy: ' + incoterms + " " + incoterms_place
        sheet.merge_cells('G16:K16')
        sheet['G16'] = 'Termin płatnosci: ' + payment_time
        sheet.merge_cells('G17:K17')
        sheet['G17'] = 'Uwaga/i: ' + str(notes)

    def add_footer(self, path: str, invoice_len: int, spec_len: int):
        workbook = load_workbook(path)

        sheet = workbook['Faktura']
        sheet['A{}'.format(invoice_len + 3)] = 'Stawkę Vat 0% zastosowano na podstawie'
        sheet['A{}'.format(invoice_len + 4)] = 'Art. 41 Ustawy o podatku od towarów i'
        sheet['A{}'.format(invoice_len + 5)] = 'usług z dnia 11 marca 2004r.'
        sheet['A{}'.format(invoice_len + 8)] = 'Sprzedający'
        sheet.merge_cells('A{}:D{}'.format(invoice_len + 9, invoice_len + 12))
        sheet['F{}'.format(invoice_len + 3)] = 'Towar od 1 do {} pozycji pochodzenia - niemieckiego'.format(
            invoice_len - 20)
        sheet['F{}'.format(invoice_len + 4)] = 'Waga towaru netto: {} kg'
        sheet['F{}'.format(invoice_len + 5)] = 'Waga towaru na {} paletach brutto: {} kg'
        sheet['F{}'.format(invoice_len + 8)] = 'Kupujacy'
        sheet.merge_cells('F{}:K{}'.format(invoice_len + 9, invoice_len + 12))

        sheet = workbook['Specyfikacja']
        sheet['A{}'.format(spec_len + 3)] = 'Stawkę Vat 0% zastosowano na podstawie'
        sheet['A{}'.format(spec_len + 4)] = 'Art. 41 Ustawy o podatku od towarów i'
        sheet['A{}'.format(spec_len + 5)] = 'usług z dnia 11 marca 2004r.'
        sheet['A{}'.format(spec_len + 8)] = 'Sprzedający'
        sheet.merge_cells('A{}:D{}'.format(spec_len + 9, spec_len + 12))
        sheet['F{}'.format(spec_len + 3)] = 'Towar od 1 do {} pozycji pochodzenia - niemieckiego'.format(spec_len - 19)
        sheet['F{}'.format(spec_len + 4)] = 'Waga towaru netto: {} kg'
        sheet['F{}'.format(spec_len + 5)] = 'Waga towaru na {} paletach brutto: {} kg'
        sheet['F{}'.format(spec_len + 8)] = 'Kupujacy'
        sheet.merge_cells('F{}:K{}'.format(spec_len + 9, spec_len + 12))

        workbook.save(path)

    def get_database_dataframe(self):
        # Query the data from the Django model
        queryset = Products.objects.all().values()

        # Convert QuerySet to DataFrame
        df = pd.DataFrame.from_records(queryset)

        return df

    def create_general_table(self, text):
        pattern = r"(\d+)\s+(\d+)\s+(\d+(?:/\d+)?)\s+([A-ZĄ-Ž\s.,1()]+(?:\s[A-Za-z0-9]+)?)\s+([A-Z]{1,3})\s+([A-Z]{2})\s+(\d+)\s+(\d{1,3}(?:\s\d{3})?,\d{2})\s+(\d{1,3}(?:\s\d{3})?,\d{2})"
        total = r"(\d+)\s([a-zA-Z]+)\s([A-Za-z.,]+)\s(EUR)\s([\d ,]+)"

        table = re.findall(pattern, text)
        total = re.findall(total, text)
        columns = ['LP.', 'Barcode', 'Kod', 'Nazwa Oryginalna', 'jm.', 'Kraj', 'Ilosc', 'Cena za szt Euro',
                   'Wartosc w EUR']
        data = [list(item) for item in table]
        df = pd.DataFrame(data, columns=columns)
        df['LP.'] = df['LP.'].astype(int)
        df['Wartosc w EUR'] = df['Wartosc w EUR'].apply(lambda x: x.replace(' ', ''))
        df['Wartosc w EUR'] = df['Wartosc w EUR'].apply(lambda x: x.replace(',', '.'))
        df['Kod'] = pd.to_numeric(df['Kod'], errors='coerce')
        df['Ilosc'] = pd.to_numeric(df['Ilosc'], errors='coerce')
        df['Wartosc w EUR'] = pd.to_numeric(df['Wartosc w EUR'], errors='coerce')

        database = self.get_database_dataframe()
        database['index'] = pd.to_numeric(database['index'], errors='coerce')
        filtered = df.merge(database, left_on='Kod', right_on='index', how='inner')

        filtered = df.merge(filtered, how='outer')
        filtered = filtered.drop(['Barcode', 'index', 'cena'], axis=1)

        filtered.insert(0, 'Podatek %', '0%')
        filtered.insert(0, 'Wartosc', filtered['Wartosc w EUR'])

        col_list = ['LP.', 'Kod', 'Nazwa Oryginalna', 'nazwa', 'kodcelny', 'Ilosc', 'jm.', 'Cena za szt Euro',
                    'Wartosc w EUR', 'Podatek %', 'Wartosc', 'waga', 'Kraj', ]
        filtered = filtered.reindex(columns=col_list)
        filtered.loc[filtered['jm.'] == 'VNT', 'jm.'] = 'szt'
        filtered['LP.'] = pd.to_numeric(filtered['LP.'])
        filtered.sort_values(by=['LP.'])

        #Test for missing values
        expected_sequence = pd.Series(range(df['LP.'].min(), df['LP.'].max() + 1))
        missing_values = expected_sequence[~expected_sequence.isin(df['LP.'])].to_list()
        total_quantity = sum(filtered['Ilosc'])
        expected_quantity = int(total[0][0])
        sum_of_eur = float(sum(filtered['Wartosc w EUR']))
        expected_sum_of_eur = float(total[0][4].replace(' ', '').replace(',', '.'))

        if expected_quantity > total_quantity:
            print("Expected quantity: {}\nFound: {}".format(expected_quantity, total_quantity))

        if expected_sum_of_eur > sum_of_eur:
            print("Expected amount: {}\nFound: {}".format(expected_sum_of_eur, sum_of_eur))

        if len(missing_values) > 0:
            print('Missing: {} positions: {}'.format(len(missing_values), missing_values))

        return {
            "table": filtered.to_dict(),
            "expected_quantity": expected_quantity,
            "total_quantity": total_quantity,
            "expected_sum_of_eur": expected_sum_of_eur,
            "sum_of_eur": sum_of_eur,
            "missing_values": missing_values,
        }

    def create_data_frames_and_save_to_excel_document(self, data_frame, excel_path):
        faktura = data_frame.loc[:, 'LP.':'Wartosc']
        quantity_sum = faktura['Ilosc'].sum()
        eur_amount_sum = faktura['Wartosc w EUR'].sum()

        faktura.loc[len(faktura)] = ['', '', '', '', 'Łącznie', quantity_sum, '', 'Suma:', eur_amount_sum, '', '']
        faktura.loc[len(faktura)] = [self.slownie(eur_amount_sum), '', '', '', '', '', '', '', '', '', '']

        specyfikacja = pd.DataFrame()

        specyfikacja.insert(0, 'LP.', data_frame['LP.'])
        specyfikacja.insert(1, 'Nazwa Oryginalna', data_frame['Nazwa Oryginalna'])
        specyfikacja.insert(2, 'Nazwa PL', data_frame['nazwa'])
        specyfikacja.insert(3, 'Kod Celny', data_frame['kodcelny'])
        specyfikacja.insert(4, 'Ilosc', data_frame['Ilosc'])
        specyfikacja.insert(5, 'Waga netto szt', data_frame['waga'])
        specyfikacja.insert(6, 'Waga netto', data_frame['waga'] * pd.to_numeric(data_frame['Ilosc']))
        specyfikacja.insert(7, 'Waga brutto', '')

        weight_sum = specyfikacja['Waga netto'].sum()

        specyfikacja.loc[len(specyfikacja)] = ['', '', '', 'Łącznie', quantity_sum, 'Łączna waga:',
                                               str(weight_sum) + 'kg', '']

        with pd.ExcelWriter(excel_path, mode="a", engine="openpyxl", if_sheet_exists='overlay') as excel_document:
            faktura.to_excel(excel_document, sheet_name="Faktura", index=False, startrow=18)
            specyfikacja.to_excel(excel_document, sheet_name="Specyfikacja", index=False, startrow=18)

        workbook = load_workbook(excel_path)
        faktura_sheet = workbook["Faktura"]
        specyfikacja_sheet = workbook["Specyfikacja"]

        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=False, showColumnStripes=False)

        faktura_table = Table(displayName="FakturaTable", ref=f"A19:K{len(faktura) + 18}")
        faktura_table.tableStyleInfo = style
        faktura_sheet.add_table(faktura_table)

        # Add the Specyfikacja table
        specyfikacja_table = Table(displayName="SpecyfikacjaTable", ref=f"A19:H{len(specyfikacja) + 18}")
        specyfikacja_table.tableStyleInfo = style
        specyfikacja_sheet.add_table(specyfikacja_table)

        # Save the workbook
        workbook.save(excel_path)

        return len(faktura) + 18, len(specyfikacja) + 18

    def extract_text_from_pdf(self, pdf):
        text = ""
        with pdfplumber.open(pdf) as pdf:
            for page in pdf.pages:
                text += page.extract_text()
        return text
    def slownie(self, liczba, skala: str = 'krótka', jeden: bool = True):
        '''
        @author Marek Madejski
        Zamienia liczbę na zapis słowny w języku polskim.
        Obsługuje liczby w zakresie do 10^66-1 dla długiej skali oraz 10^36-1 dla krótkiej skali.
        Możliwe pominięcie słowa "jeden" przy potęgach tysiąca.
        '''

        prefix = "slownie: "
        suffix = " " + str(liczba).split(".")[1][:2] + '/100 '
        liczba = int(liczba)
        if (skala == 'długa' and abs(liczba) >= 10 ** 66) or (skala == 'krótka' and abs(liczba) >= 10 ** 36):
            raise ValueError('Zbyt duża liczba.')

        jedności = ('', 'jeden', 'dwa', 'trzy', 'cztery', 'pięć', 'sześć', 'siedem', 'osiem', 'dziewięć')
        naście = (
            '', 'jedenaście', 'dwanaście', 'trzynaście', 'czternaście', 'piętnaście', 'szesnaście', 'siedemnaście',
            'osiemnaście', 'dziewiętnaście')
        dziesiątki = (
            '', 'dziesięć', 'dwadzieścia', 'trzydzieści', 'czterdzieści', 'pięćdziesiąt', 'sześćdziesiąt',
            'siedemdziesiąt',
            'osiemdziesiąt', 'dziewięćdziesiąt')
        setki = (
            '', 'sto', 'dwieście', 'trzysta', 'czterysta', 'pięćset', 'sześćset', 'siedemset', 'osiemset',
            'dziewięćset')

        grupy = [('', '', ''), ('tysiąc', 'tysiące', 'tysięcy')]

        przedrostki = ('mi', 'bi', 'try', 'kwadry', 'kwinty', 'seksty', 'septy', 'okty', 'nony', 'decy')
        for p in przedrostki:
            grupy.append((f'{p}lion', f'{p}liony', f'{p}lionów'))
            if skala == 'długa':
                grupy.append((f'{p}liard', f'{p}liardy', f'{p}liardów'))

        if liczba == 0:
            return 'zero'

        słowa = []
        znak = ''
        if liczba < 0:
            znak = 'minus'
            liczba = -liczba
        g = 0
        while liczba != 0:
            s = liczba % 1_000 // 100
            d = liczba % 100 // 10
            j = liczba % 10
            liczba //= 1_000

            if s == d == j == 0:
                g += 1
                continue

            if d == 1 and j > 0:
                n = j
                d = j = 0
            else:
                n = 0

            if j == 1 and s + d + n == 0:
                forma = 0
            elif 2 <= j <= 4:
                forma = 1
            else:
                forma = 2

            słowa = [setki[s], dziesiątki[d], naście[n], jedności[j] if jeden or g == 0 else '',
                     grupy[g][forma]] + słowa
            g += 1

        słowa.insert(0, znak)
        return prefix + ' '.join(s for s in słowa if s) + suffix

    if __name__ == "__main__":
        main()

# df.to_excel('output.xlsx', index=False,sheet_name='FAKTURA',startrow=10)
# filtered.to_excel('filtered.xlsx',index=False, sheet_name='Faktura')
